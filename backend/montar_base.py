import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIR = Path(__file__).parent.parent / 'data'
ENTRADA = BASE_DIR / 'base'
SAIDA = BASE_DIR / 'base_formatada'
COORD_FILE = BASE_DIR / 'coordenadores' / 'Coordenadores.xlsx'

ENTRADA.mkdir(parents=True, exist_ok=True)
SAIDA.mkdir(parents=True, exist_ok=True)

# Arquivo mais recente em data/base
arquivos = sorted(ENTRADA.glob('*.xlsx'), key=os.path.getmtime, reverse=True)
if not arquivos:
    raise FileNotFoundError("Nenhum arquivo .xlsx encontrado em data/base")
arquivo = arquivos[0]
print(f"Processando: {arquivo.name}")

# Unir as duas abas
xl = pd.ExcelFile(arquivo)
sheet_pendente = next(s for s in xl.sheet_names if 'Ok' not in s)
sheet_ok = next(s for s in xl.sheet_names if 'Ok' in s)
df = pd.concat(
    [pd.read_excel(xl, sheet_name=sheet_pendente),
     pd.read_excel(xl, sheet_name=sheet_ok)],
    ignore_index=True,
)

# Coluna Status via de/para de Estagio
DE_PARA = {
    '19 - Arquivo não recebido (Arquivo não transmitido)': 'Não transmitido',
    '27 - Reaberto pela controladoria (Arquivo não transmitido)': 'Não transmitido',
    '12 - Arquivo Transmitido (Arquivo transmitido)': 'Transmitido',
    '13 - Arquivo Transmitido Sem Movimento (Arquivo transmitido)': 'Transmitido',
    '33 - IE Suspensa/Inapta (Arquivo não transmitido)': 'Transmitido',
    '35 - Arquivo Transmitido - Ex-Cliente (Arquivo transmitido)': 'Transmitido',
}
df['Status'] = df['Estagio'].map(DE_PARA)

# DataAlteracaoEstagio: somente data, sem hora
df['DataAlteracaoEstagio'] = pd.to_datetime(df['DataAlteracaoEstagio']).dt.normalize()

CORTE = pd.Timestamp('2026-06-01')

# Não transmitido → em branco
df.loc[df['Status'] == 'Não transmitido', 'DataAlteracaoEstagio'] = pd.NaT

# Transmitido com data anterior ao corte → 01/06/2026
mask_ok = df['Status'] == 'Transmitido'
df.loc[mask_ok & (df['DataAlteracaoEstagio'] < CORTE), 'DataAlteracaoEstagio'] = CORTE

# Coluna Tributacao via de/para de RegimeTributario
DE_PARA_TRIBUTACAO = {
    'Federal - Imune': 'Imune',
    'Federal - Lucro Presumido': 'Lucro Presumido',
    'Federal - L Real -Trimestral': 'Lucro Real',
    'Federal - L.Real - Mensal': 'Lucro Real',
    'Federal - SN': 'Simples Nacional',
    'Federal - Lucro Real - Anual': 'Lucro Real',
}
df['Tributacao'] = df['RegimeTributario'].map(DE_PARA_TRIBUTACAO)

# Coordenador(a) a partir de ResponsavelPadrao
coord = pd.read_excel(COORD_FILE)
coord_map = coord.set_index('NOME EXIBICAO')['Coordenador'].to_dict()
df['Coordenador(a)'] = df['ResponsavelPadrao'].map(coord_map)

# Salvar
nome_saida = SAIDA / 'Base.xlsx'
df.to_excel(nome_saida, index=False, engine='openpyxl')

# Pós-processamento: formatar data e converter em tabela Excel
wb = load_workbook(nome_saida)
ws = wb.active
header = [cell.value for cell in ws[1]]

# Formatar DataAlteracaoEstagio como DD/MM/AAAA
col_idx = header.index('DataAlteracaoEstagio') + 1
for (cell,) in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
    if cell.value is not None:
        cell.number_format = 'DD/MM/YYYY'

# Converter em tabela Excel
ultima_col = get_column_letter(len(header))
ultima_linha = ws.max_row
ref = f"A1:{ultima_col}{ultima_linha}"
tabela = Table(displayName="Base", ref=ref)
tabela.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(tabela)

wb.save(nome_saida)
print(f"Concluído! Arquivo salvo em: {nome_saida}")
